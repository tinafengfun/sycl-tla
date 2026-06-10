#################################################################################################
# Copyright (C) 2026 Intel Corporation, All rights reserved.
# SPDX-License-Identifier: BSD-3-Clause
#################################################################################################

from pathlib import Path

from .schemas import SCHEMA_VERSION
from .utils import now_iso


ALI_PROVIDER_COLUMNS = {
    "bf16": {
        "oneMKL(BLAS)": 4,
        "oneDNN(matmul)": 5,
        "PyTorch->oneDNN": 6,
        "Sycl_TLA(00_base)": 7,
        "Sycl_TLA(00_padded)": 8,
        "Sycl_TLA(00_sycl_q)": 9,
        "Sycl_TLA(03_streamk)": 10,
        "Sycl_TLA(03_dp)": 11,
    },
    "f16": {
        "oneMKL(BLAS)": 15,
        "oneDNN(matmul)": 16,
        "PyTorch->oneDNN": 17,
        "SYCL-TLA(XeTLA)": 18,
    },
    "int8": {
        "oneMKL(BLAS)": 19,
        "oneDNN(matmul)": 20,
        "PyTorch->oneDNN": 21,
        "SYCL-TLA(XeTLA)": 22,
    },
}

ALI_REFERENCE_PROVIDERS = {
    "bf16": [
        "Sycl_TLA(00_base)",
        "Sycl_TLA(00_padded)",
        "Sycl_TLA(00_sycl_q)",
        "Sycl_TLA(03_streamk)",
        "Sycl_TLA(03_dp)",
    ],
    "f16": ["SYCL-TLA(XeTLA)"],
    "int8": ["SYCL-TLA(XeTLA)"],
}


def _as_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _shape_id(layout, dtype, m, n, k):
    return f"{layout}_{dtype}_{m}_{n}_{k}"


def _best_provider(metrics, provider_names=None):
    source = metrics
    if provider_names is not None:
        source = {name: metrics.get(name) for name in provider_names}
    available = [(name, value) for name, value in source.items() if value is not None]
    if not available:
        return "", None
    provider, value = max(available, key=lambda item: item[1])
    return provider, value


def build_ali_gemm_docs(workbook_path, layouts=("rcr", "rrr"), supported_dtypes=("bf16", "f16")):
    from openpyxl import load_workbook

    workbook = load_workbook(Path(workbook_path), data_only=True, read_only=True)
    entries = []
    shapes = []
    skipped = []
    seen_shapes = set()

    for sheet in workbook.worksheets:
        header_row_index = None
        header_row = None
        for row_index, candidate_header in enumerate(
            sheet.iter_rows(min_row=1, max_row=5, values_only=True),
            start=1,
        ):
            first = candidate_header[0] if candidate_header else None
            second = candidate_header[1] if len(candidate_header) > 1 else None
            if first == "M" or second == "M":
                header_row_index = row_index
                header_row = candidate_header
                break
        if header_row_index is None or header_row is None:
            continue
        offset = 0 if header_row and header_row[0] == "M" else 1
        for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            m = row[0 + offset] if len(row) > 0 + offset else None
            n = row[1 + offset] if len(row) > 1 + offset else None
            k = row[2 + offset] if len(row) > 2 + offset else None
            shape_type = row[3 + offset] if len(row) > 3 + offset else ""
            if m in (None, "") or n in (None, "") or k in (None, ""):
                continue
            for dtype, provider_columns in ALI_PROVIDER_COLUMNS.items():
                metrics = {
                    provider: _as_float(row[index + offset]) if index + offset < len(row) else None
                    for provider, index in provider_columns.items()
                }
                provider, best_tflops = _best_provider(metrics, ALI_REFERENCE_PROVIDERS.get(dtype))
                if best_tflops is None:
                    continue
                for layout in layouts:
                    entry = {
                        "sheet": sheet.title,
                        "layout": layout,
                        "dtype_a": dtype,
                        "dtype_b": dtype,
                        "dtype_c": "f32",
                        "dtype_d": "f32",
                        "dtype_acc": "f32",
                        "m": int(m),
                        "n": int(n),
                        "k": int(k),
                        "batch_count": 1,
                        "shape_type": str(shape_type or "").strip(),
                        "shape_id": _shape_id(layout, dtype, int(m), int(n), int(k)),
                        "providers": metrics,
                        "reference_provider": provider,
                        "reference_tflops": best_tflops,
                        "supported": dtype in supported_dtypes,
                    }
                    entries.append(entry)
                    if entry["supported"]:
                        shape_key = (
                            entry["shape_id"],
                            entry["dtype_a"],
                            entry["m"],
                            entry["n"],
                            entry["k"],
                        )
                        if shape_key not in seen_shapes:
                            seen_shapes.add(shape_key)
                            shapes.append(
                                {
                                    "shape_id": entry["shape_id"],
                                    "layout": layout,
                                    "dtype_a": dtype,
                                    "dtype_b": dtype,
                                    "dtype_c": "f32",
                                    "dtype_d": "f32",
                                    "dtype_acc": "f32",
                                    "m": int(m),
                                    "n": int(n),
                                    "k": int(k),
                                    "batch_count": 1,
                                    "runtime_defaults": {},
                                    "tags": [entry["shape_type"].replace(" ", "_").lower()] if entry["shape_type"] else [],
                                }
                            )
                else:
                    skipped.append(
                        {
                            "sheet": sheet.title,
                            "dtype": dtype,
                            "m": int(m),
                            "n": int(n),
                            "k": int(k),
                            "reason": "dtype_not_supported_by_profiler",
                        }
                    )

    shapes_doc = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": now_iso(),
        "shape_set_id": "ali_gemm_perf_supported",
        "source": str(Path(workbook_path)),
        "shapes": shapes,
    }
    reference_doc = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": now_iso(),
        "dataset_id": "ali_gemm_perf_reference",
        "source": str(Path(workbook_path)),
        "supported_dtypes": list(supported_dtypes),
        "entries": entries,
        "skipped_entries": skipped,
    }
    return shapes_doc, reference_doc
